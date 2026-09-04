"""
Customer management routes blueprint.
Handles Customer CRUD operations, statement generation, file upload parsing,
and CSV/Excel importing and exporting.
"""
from datetime import datetime, date, timedelta
import csv
import io
import os
import uuid
from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app, send_file, jsonify
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook

from database import db
from models.customer import Customer
from models.sale import Sale
from models.payment import Payment
from models.instalment_plan import InstalmentPlan
from models.instalment_schedule import InstalmentSchedule
from models.audit_log import AuditLog
from forms.customer import CustomerForm
from utils.auth_decorators import role_required

customers_bp = Blueprint('customers', __name__, url_prefix='/customers')

# --- Helper function for file uploads ---

def save_profile_photo(file_data):
    """
    Saves profile photos with unique secure UUID filenames.
    Creates path if missing. Returns relative web path.
    """
    if not file_data or not file_data.filename:
        return None
        
    # Build target directory path
    upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'profile_photos')
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir, exist_ok=True)
        
    ext = file_data.filename.rsplit('.', 1)[-1].lower()
    unique_name = f"{uuid.uuid4().hex}.{ext}"
    file_path = os.path.join(upload_dir, unique_name)
    
    file_data.save(file_path)
    return f"uploads/profile_photos/{unique_name}"

# --- Helper function for code generation ---

def generate_customer_code():
    """
    Calculates next available CUSTXXXX format sequence.
    """
    last_customer = Customer.query.order_by(Customer.customer_code.desc()).first()
    if not last_customer:
        return "CUST0001"
    
    code = last_customer.customer_code
    try:
        num = int(code.replace("CUST", ""))
        new_num = num + 1
        return f"CUST{new_num:04d}"
    except ValueError:
        return f"CUST{uuid.uuid4().hex[:4].upper()}"


# --- Routes ---

@customers_bp.route('/')
@login_required
def index():
    """
    Renders customer list page with filter queries.
    """
    filter_type = request.args.get('filter', '').strip().lower()
    query = Customer.query.filter(Customer.deleted_at == None)
    
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = today.replace(day=1)
    
    if filter_type == 'active':
        query = query.filter(Customer.status == 'active')
    elif filter_type == 'inactive':
        query = query.filter(Customer.status == 'inactive')
    elif filter_type == 'today':
        query = query.filter(Customer.created_at >= today)
    elif filter_type == 'this_month':
        query = query.filter(Customer.created_at >= month_start)
    elif filter_type == 'pending_payments':
        # Subquery for customers with active/overdue instalments
        query = query.join(InstalmentPlan).filter(InstalmentPlan.status.in_(['active', 'overdue']))
    elif filter_type == 'completed_instalments':
        # Customers who have completed plans and no active plans
        completed_sub = db.session.query(InstalmentPlan.customer_id).filter(InstalmentPlan.status == 'completed').subquery()
        active_sub = db.session.query(InstalmentPlan.customer_id).filter(InstalmentPlan.status.in_(['active', 'overdue'])).subquery()
        query = query.filter(Customer.id.in_(completed_sub)).filter(~Customer.id.in_(active_sub))

    customers = query.order_by(Customer.customer_code.desc()).all()
    return render_template('customers/list.html', customers=customers, current_filter=filter_type)


@customers_bp.route('/register', methods=['GET', 'POST'])
@login_required
def register():
    """
    Registers a new credit customer.
    """
    form = CustomerForm()
    if form.validate_on_submit():
        photo_path = None
        if form.profile_photo.data:
            photo_path = save_profile_photo(form.profile_photo.data)
            
        code = generate_customer_code()
        
        customer = Customer(
            customer_code=code,
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            full_name=f"{form.first_name.data} {form.last_name.data}",
            phone_number=form.phone_number.data,
            alternate_phone=form.alternate_phone.data,
            email=form.email.data,
            address_line1=form.address_line1.data,
            address_line2=form.address_line2.data,
            city=form.city.data,
            state=form.state.data,
            postal_code=form.postal_code.data,
            country=form.country.data,
            date_of_birth=form.date_of_birth.data,
            gender=form.gender.data,
            occupation=form.occupation.data,
            aadhaar_number=form.aadhaar_number.data,
            profile_photo=photo_path,
            notes=form.notes.data,
            created_by=current_user.id,
            status='active'
        )
        
        db.session.add(customer)
        
        # Log Audit Trail
        log = AuditLog(
            user_id=current_user.id,
            username=current_user.username,
            action=f"customer_create: {code}",
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f"Customer {customer.full_name} registered successfully with code {code}.", "success")
        return redirect(url_for('customers.index'))
        
    return render_template('customers/register.html', form=form)


@customers_bp.route('/edit/<string:id>', methods=['GET', 'POST'])
@login_required
@role_required(['Super Admin', 'Admin'])
def edit(id):
    """
    Edits existing customer details.
    Staff restricted.
    """
    customer = Customer.query.filter_by(id=id, deleted_at=None).first_or_404()
    form = CustomerForm(customer_id=customer.id, obj=customer)
    
    if form.validate_on_submit():
        if form.profile_photo.data:
            # Drop old file if exists to clean disk storage
            if customer.profile_photo:
                old_path = os.path.join(current_app.root_path, 'static', customer.profile_photo)
                if os.path.exists(old_path):
                    try:
                        os.remove(old_path)
                    except OSError:
                        pass
            customer.profile_photo = save_profile_photo(form.profile_photo.data)
            
        customer.first_name = form.first_name.data
        customer.last_name = form.last_name.data
        customer.full_name = f"{form.first_name.data} {form.last_name.data}"
        customer.phone_number = form.phone_number.data
        customer.alternate_phone = form.alternate_phone.data
        customer.email = form.email.data
        customer.address_line1 = form.address_line1.data
        customer.address_line2 = form.address_line2.data
        customer.city = form.city.data
        customer.state = form.state.data
        customer.postal_code = form.postal_code.data
        customer.country = form.country.data
        customer.date_of_birth = form.date_of_birth.data
        customer.gender = form.gender.data
        customer.occupation = form.occupation.data
        customer.aadhaar_number = form.aadhaar_number.data
        customer.notes = form.notes.data
        customer.updated_by = current_user.id
        
        # Log Audit Trail
        log = AuditLog(
            user_id=current_user.id,
            username=current_user.username,
            action=f"customer_update: {customer.customer_code}",
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f"Customer {customer.full_name} profile updated successfully.", "success")
        return redirect(url_for('customers.profile', id=customer.id))
        
    return render_template('customers/edit.html', form=form, customer=customer)


@customers_bp.route('/profile/<string:id>')
@login_required
def profile(id):
    """
    Renders detailed Customer profile, purchases history, and timeline.
    """
    customer = Customer.query.filter_by(id=id, deleted_at=None).first_or_404()
    
    # 1. Timeline Chronology Events Builder
    timeline = []
    
    # Customer registration event
    timeline.append({
        'date': customer.created_at,
        'title': 'Customer Account Created',
        'desc': f"Customer profile registered with code {customer.customer_code}.",
        'icon': 'bi-person-check',
        'color': 'primary'
    })
    
    # Sales transactions events
    for sale in customer.sales:
        timeline.append({
            'date': sale.sale_date,
            'title': f"Purchased {sale.product.name}",
            'desc': f"Invoice {sale.invoice_number} created. Total value: ₹{sale.total_amount:,.2f} with down-payment: ₹{sale.down_payment:,.2f}.",
            'icon': 'bi-cart-check',
            'color': 'success'
        })
        
        # Instalment plan completions events
        if sale.instalment_plan and sale.instalment_plan.status == 'completed':
            timeline.append({
                'date': sale.instalment_plan.updated_at,
                'title': f"Instalment Plan Completed",
                'desc': f"Repayment plan for Invoice {sale.invoice_number} was fully cleared and closed.",
                'icon': 'bi-award',
                'color': 'info'
            })
            
    # Payments collections events
    for pay in customer.payments:
        timeline.append({
            'date': pay.payment_date,
            'title': f"Payment Received: ₹{pay.payment_amount:,.2f}",
            'desc': f"Receipt ID: REC-{pay.id} recorded against Invoice {pay.sale.invoice_number}.",
            'icon': 'bi-cash-coin',
            'color': 'warning'
        })
        
    # Sort events by date descending
    timeline.sort(key=lambda x: x['date'], reverse=True)
    
    # Log Audit Trail
    log = AuditLog(
        user_id=current_user.id,
        username=current_user.username,
        action=f"customer_view_profile: {customer.customer_code}",
        ip_address=request.remote_addr,
        user_agent=request.user_agent.string
    )
    db.session.add(log)
    db.session.commit()
    
    # Summary items counts
    active_insts_count = sum(1 for plan in customer.instalment_plans if plan.status in ['active', 'overdue'])
    completed_insts_count = sum(1 for plan in customer.instalment_plans if plan.status == 'completed')
    last_payment = Payment.query.filter_by(customer_id=customer.id, status='paid').order_by(Payment.payment_date.desc()).first()
    
    return render_template(
        'customers/profile.html', 
        customer=customer, 
        timeline=timeline,
        active_insts_count=active_insts_count,
        completed_insts_count=completed_insts_count,
        last_payment_date=last_payment.payment_date if last_payment else None
    )


@customers_bp.route('/delete/<string:id>', methods=['POST'])
@login_required
@role_required(['Super Admin', 'Admin'])
def delete(id):
    """
    Soft deletes customer record.
    Restricts delete if there's any active outstanding instalment balance.
    """
    customer = Customer.query.filter_by(id=id, deleted_at=None).first_or_404()
    
    if customer.outstanding_balance > 0:
        flash(f"Cannot delete customer {customer.full_name}. They have an active outstanding balance of ₹{customer.outstanding_balance:,.2f}.", "danger")
        return redirect(url_for('customers.index'))
        
    customer.deleted_at = datetime.utcnow()
    
    # Log Audit Trail
    log = AuditLog(
        user_id=current_user.id,
        username=current_user.username,
        action=f"customer_delete: {customer.customer_code}",
        ip_address=request.remote_addr,
        user_agent=request.user_agent.string
    )
    db.session.add(log)
    db.session.commit()
    
    flash(f"Customer {customer.full_name} has been soft-deleted successfully.", "success")
    return redirect(url_for('customers.index'))


@customers_bp.route('/statement/<string:id>')
@login_required
def statement(id):
    """
    Renders print-friendly invoice statements.
    """
    customer = Customer.query.filter_by(id=id, deleted_at=None).first_or_404()
    
    # Log Audit Trail
    log = AuditLog(
        user_id=current_user.id,
        username=current_user.username,
        action=f"customer_statement_generate: {customer.customer_code}",
        ip_address=request.remote_addr,
        user_agent=request.user_agent.string
    )
    db.session.add(log)
    db.session.commit()
    
    return render_template('customers/statement.html', customer=customer, today=date.today())


@customers_bp.route('/export/<string:format_type>')
@login_required
@role_required(['Super Admin', 'Admin'])
def export_data(format_type):
    """
    Exports Customer database table as Excel (.xlsx) or CSV.
    """
    customers = Customer.query.filter(Customer.deleted_at == None).all()
    
    # Audit log
    log = AuditLog(
        user_id=current_user.id,
        username=current_user.username,
        action=f"customer_export: {format_type}",
        ip_address=request.remote_addr,
        user_agent=request.user_agent.string
    )
    db.session.add(log)
    db.session.commit()

    if format_type == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Customer Code', 'Full Name', 'Phone', 'Email', 'City', 'State', 'Outstanding Balance', 'Status', 'Date Joined'])
        
        for c in customers:
            writer.writerow([
                c.customer_code, c.full_name, c.phone_number, c.email or '',
                c.city, c.state, c.outstanding_balance, c.status, c.created_at.strftime('%Y-%m-%d')
            ])
            
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"customers_export_{datetime.now().strftime('%Y%m%d')}.csv"
        )
        
    elif format_type == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers List"
        
        # Headers
        ws.append(['Customer Code', 'First Name', 'Last Name', 'Phone', 'Email', 'Address', 'City', 'State', 'Postal Code', 'Outstanding Balance', 'Status', 'Created At'])
        
        for c in customers:
            ws.append([
                c.customer_code, c.first_name, c.last_name, c.phone_number, c.email or '',
                c.address_line1, c.city, c.state, c.postal_code, c.outstanding_balance, c.status, c.created_at.strftime('%Y-%m-%d')
            ])
            
        out_stream = io.BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        return send_file(
            out_stream,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"customers_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
    flash("Invalid export format requested.", "danger")
    return redirect(url_for('customers.index'))


@customers_bp.route('/import', methods=['POST'])
@login_required
@role_required(['Super Admin', 'Admin'])
def import_data():
    """
    Imports customer list from CSV or Excel file formats.
    """
    file = request.files.get('import_file')
    if not file or not file.filename:
        flash("Please upload a file to import.", "danger")
        return redirect(url_for('customers.index'))
        
    ext = file.filename.rsplit('.', 1)[-1].lower()
    import_count = 0
    
    try:
        if ext == 'csv':
            stream = io.StringIO(file.stream.read().decode("utf-8"), newline=None)
            reader = csv.reader(stream)
            header = next(reader, None) # Skip header
            
            for row in reader:
                if len(row) < 7:
                    continue # Bad format
                
                # Check duplicate phone
                phone = row[3].strip()
                if Customer.query.filter_by(phone_number=phone, deleted_at=None).first():
                    continue # Skip duplicate
                    
                code = generate_customer_code()
                cust = Customer(
                    customer_code=code,
                    first_name=row[1].strip(),
                    last_name=row[2].strip(),
                    full_name=f"{row[1].strip()} {row[2].strip()}",
                    phone_number=phone,
                    email=row[4].strip() if row[4] else None,
                    address_line1=row[5].strip(),
                    city=row[6].strip(),
                    state=row[7].strip() if len(row) > 7 else 'Delhi',
                    postal_code=row[8].strip() if len(row) > 8 else '110001',
                    date_of_birth=datetime.strptime('1990-01-01', '%Y-%m-%d').date(), # Default
                    gender='Male', # Default
                    created_by=current_user.id
                )
                db.session.add(cust)
                import_count += 1
                
        elif ext in ['xls', 'xlsx']:
            wb = load_workbook(file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                # Loop skip header
                for row in rows[1:]:
                    if not row or len(row) < 5 or not row[3]:
                        continue
                        
                    phone = str(row[3]).strip()
                    if Customer.query.filter_by(phone_number=phone, deleted_at=None).first():
                        continue
                        
                    code = generate_customer_code()
                    cust = Customer(
                        customer_code=code,
                        first_name=str(row[1]).strip(),
                        last_name=str(row[2]).strip(),
                        full_name=f"{str(row[1]).strip()} {str(row[2]).strip()}",
                        phone_number=phone,
                        email=str(row[4]).strip() if row[4] else None,
                        address_line1=str(row[5]).strip() if len(row) > 5 else 'Address Line',
                        city=str(row[6]).strip() if len(row) > 6 else 'City',
                        state=str(row[7]).strip() if len(row) > 7 else 'State',
                        postal_code=str(row[8]).strip() if len(row) > 8 else '000000',
                        date_of_birth=datetime.strptime('1990-01-01', '%Y-%m-%d').date(),
                        gender='Male',
                        created_by=current_user.id
                    )
                    db.session.add(cust)
                    import_count += 1
                    
        else:
            flash("Invalid format. Only CSV or XLSX files are allowed.", "danger")
            return redirect(url_for('customers.index'))
            
        if import_count > 0:
            # Log Audit Trail
            log = AuditLog(
                user_id=current_user.id,
                username=current_user.username,
                action=f"customer_import: {import_count} entries",
                ip_address=request.remote_addr,
                user_agent=request.user_agent.string
            )
            db.session.add(log)
            db.session.commit()
            flash(f"Imported {import_count} customer records successfully.", "success")
        else:
            flash("No new customer records imported (all records matched existing phone numbers).", "warning")
            
    except Exception as e:
        db.session.rollback()
        flash(f"Import failed with error: {str(e)}", "danger")
        
    return redirect(url_for('customers.index'))
