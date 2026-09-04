"""
Product management routes blueprint.
Handles Product CRUD, barcode printing, and spreadsheet imports/exports.
"""
from datetime import datetime
import csv
import io
import os
import uuid
import decimal
from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook

from database import db
from models.product import Product
from models.category import Category
from models.brand import Brand
from models.inventory_movement import InventoryMovement
from models.audit_log import AuditLog
from forms.product import ProductForm
from utils.auth_decorators import role_required

products_bp = Blueprint('products', __name__, url_prefix='/products')

# --- Helpers ---

def save_product_image(file_data):
    """
    Saves product photos with unique secure UUID filenames.
    """
    if not file_data or not file_data.filename:
        return None
        
    upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'products')
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir, exist_ok=True)
        
    ext = file_data.filename.rsplit('.', 1)[-1].lower()
    unique_name = f"{uuid.uuid4().hex}.{ext}"
    file_path = os.path.join(upload_dir, unique_name)
    
    file_data.save(file_path)
    return f"uploads/products/{unique_name}"

def generate_product_code():
    """
    Generates next sequential product code.
    """
    last_prod = Product.query.order_by(Product.product_code.desc()).first()
    if not last_prod:
        return "PRD0001"
    
    code = last_prod.product_code
    try:
        num = int(code.replace("PRD", ""))
        return f"PRD{num + 1:04d}"
    except ValueError:
        return f"PRD{uuid.uuid4().hex[:4].upper()}"


# --- Routes ---

@products_bp.route('/')
@login_required
def index():
    """
    Renders product listings grid with filters.
    """
    filter_type = request.args.get('filter', '').strip().lower()
    cat_id = request.args.get('category', '').strip()
    brand_id = request.args.get('brand', '').strip()
    
    query = Product.query.filter(Product.deleted_at == None)
    
    # Apply category/brand filters
    if cat_id:
        query = query.filter(Product.category_id == cat_id)
    if brand_id:
        query = query.filter(Product.brand_id == brand_id)
        
    # Apply stock filters
    if filter_type == 'in_stock':
        query = query.filter(Product.current_stock > Product.minimum_stock)
    elif filter_type == 'low_stock':
        query = query.filter(Product.current_stock <= Product.minimum_stock, Product.current_stock > 0)
    elif filter_type == 'out_of_stock':
        query = query.filter(Product.current_stock <= 0)
    elif filter_type == 'overstock':
        query = query.filter(Product.current_stock > Product.maximum_stock)
    elif filter_type == 'active':
        query = query.filter(Product.status == 'active')
    elif filter_type == 'inactive':
        query = query.filter(Product.status == 'inactive')
        
    products = query.order_by(Product.product_code.desc()).all()
    categories = Category.query.filter_by(status='active').all()
    brands = Brand.query.filter_by(status='active').all()
    
    return render_template(
        'products/list.html',
        products=products,
        categories=categories,
        brands=brands,
        current_filter=filter_type,
        current_cat=cat_id,
        current_brand=brand_id
    )

@products_bp.route('/add', methods=['GET', 'POST'])
@login_required
@role_required(['Super Admin', 'Admin'])
def add():
    """
    Creates a new product record and logs opening inventory.
    """
    form = ProductForm()
    
    # Populate Category and Brand dropdown choices dynamically
    categories = Category.query.filter_by(status='active').all()
    brands = Brand.query.filter_by(status='active').all()
    
    form.category_id.choices = [('', '-- Select Category --')] + [(c.id, f"{c.name} ({c.category_code})") for c in categories]
    form.brand_id.choices = [('', '-- Select Brand --')] + [(b.id, b.brand_name) for b in brands]
    
    if form.validate_on_submit():
        image_path = None
        if form.product_image.data:
            image_path = save_product_image(form.product_image.data)
            
        code = generate_product_code()
        
        # Enforce business rule: stock values must align
        open_stock = form.opening_stock.data or 0
        
        prod = Product(
            product_code=code,
            barcode=code, # Barcode defaults to product code
            qr_code=code,
            product_name=form.product_name.data,
            category_id=form.category_id.data,
            brand_id=form.brand_id.data,
            description=form.description.data,
            purchase_price=form.purchase_price.data,
            selling_price=form.selling_price.data,
            gst_percentage=form.gst_percentage.data,
            discount_percentage=form.discount_percentage.data or 0,
            current_stock=open_stock,
            minimum_stock=form.minimum_stock.data,
            maximum_stock=form.maximum_stock.data,
            unit=form.unit.data,
            product_image=image_path,
            status='active',
            created_by=current_user.id
        )
        
        db.session.add(prod)
        db.session.flush() # Yields prod.id for movement logging
        
        # Log opening stock movement
        if open_stock > 0:
            mov = InventoryMovement(
                product_id=prod.id,
                movement_type='Stock In',
                quantity=open_stock,
                previous_stock=0,
                new_stock=open_stock,
                remarks='Opening balance initial stock.',
                created_by=current_user.id
            )
            db.session.add(mov)
            
        # Audit log entry
        log = AuditLog(
            user_id=current_user.id,
            username=current_user.username,
            action=f"product_create: {code} ({prod.product_name})",
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f"Product '{prod.product_name}' registered successfully with code {code}.", "success")
        return redirect(url_for('products.index'))
        
    return render_template('products/add.html', form=form)

@products_bp.route('/edit/<string:id>', methods=['GET', 'POST'])
@login_required
@role_required(['Super Admin', 'Admin'])
def edit(id):
    """
    Edits existing product details.
    """
    prod = Product.query.filter_by(id=id, deleted_at=None).first_or_404()
    form = ProductForm(product_id=prod.id, obj=prod)
    
    categories = Category.query.filter_by(status='active').all()
    brands = Brand.query.filter_by(status='active').all()
    form.category_id.choices = [('', '-- Select Category --')] + [(c.id, f"{c.name} ({c.category_code})") for c in categories]
    form.brand_id.choices = [('', '-- Select Brand --')] + [(b.id, b.brand_name) for b in brands]
    
    if form.validate_on_submit():
        if form.product_image.data:
            # Delete old image if exists
            if prod.product_image:
                old_path = os.path.join(current_app.root_path, 'static', prod.product_image)
                if os.path.exists(old_path):
                    try:
                        os.remove(old_path)
                    except OSError:
                        pass
            prod.product_image = save_product_image(form.product_image.data)
            
        prod.product_name = form.product_name.data
        prod.category_id = form.category_id.data
        prod.brand_id = form.brand_id.data
        prod.description = form.description.data
        prod.purchase_price = form.purchase_price.data
        prod.selling_price = form.selling_price.data
        prod.gst_percentage = form.gst_percentage.data
        prod.discount_percentage = form.discount_percentage.data or 0
        prod.minimum_stock = form.minimum_stock.data
        prod.maximum_stock = form.maximum_stock.data
        prod.unit = form.unit.data
        prod.updated_by = current_user.id
        
        # Log Audit
        log = AuditLog(
            user_id=current_user.id,
            username=current_user.username,
            action=f"product_update: {prod.product_code}",
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f"Product '{prod.product_name}' details updated successfully.", "success")
        return redirect(url_for('products.detail', id=prod.id))
        
    # Pre-select fields
    form.category_id.data = prod.category_id
    form.brand_id.data = prod.brand_id
    
    return render_template('products/edit.html', form=form, product=prod)

@products_bp.route('/detail/<string:id>')
@login_required
def detail(id):
    """
    Renders detailed product profile, barcode view, and movement timeline.
    """
    prod = Product.query.filter_by(id=id, deleted_at=None).first_or_404()
    movements = InventoryMovement.query.filter_by(product_id=id).order_by(InventoryMovement.created_at.desc()).all()
    
    # Calculate analytical stats (purchases value, total revenue generated)
    total_sales_count = 0 # Future Billing connection placeholders
    total_revenue = 0.0
    
    return render_template(
        'products/detail.html',
        product=prod,
        movements=movements,
        total_sales_count=total_sales_count,
        total_revenue=total_revenue
    )

@products_bp.route('/delete/<string:id>', methods=['POST'])
@login_required
@role_required(['Super Admin', 'Admin'])
def delete(id):
    """
    Soft deletes product record.
    """
    prod = Product.query.filter_by(id=id, deleted_at=None).first_or_404()
    prod.deleted_at = datetime.utcnow()
    
    # Audit log
    log = AuditLog(
        user_id=current_user.id,
        username=current_user.username,
        action=f"product_delete: {prod.product_code}",
        ip_address=request.remote_addr,
        user_agent=request.user_agent.string
    )
    db.session.add(log)
    db.session.commit()
    
    flash(f"Product '{prod.product_name}' has been soft-deleted successfully.", "success")
    return redirect(url_for('products.index'))

@products_bp.route('/barcode-print/<string:id>')
@login_required
def barcode_print(id):
    """
    Displays print-friendly barcode labels sheet.
    """
    prod = Product.query.filter_by(id=id, deleted_at=None).first_or_404()
    
    # Audit Log
    log = AuditLog(
        user_id=current_user.id,
        username=current_user.username,
        action=f"barcode_print: {prod.product_code}",
        ip_address=request.remote_addr,
        user_agent=request.user_agent.string
    )
    db.session.add(log)
    db.session.commit()
    
    return render_template('products/barcode_print.html', product=prod)

@products_bp.route('/export/<string:format_type>')
@login_required
@role_required(['Super Admin', 'Admin'])
def export_data(format_type):
    """
    Exports Product inventory table as Excel or CSV.
    """
    products = Product.query.filter(Product.deleted_at == None).all()

    if format_type == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Product Code', 'Product Name', 'Barcode', 'Category', 'Brand', 'Purchase Price', 'Selling Price', 'Current Stock', 'Unit', 'Status'])
        
        for p in products:
            writer.writerow([
                p.product_code, p.product_name, p.barcode, p.category.name, p.brand.brand_name,
                float(p.purchase_price), float(p.selling_price), p.current_stock, p.unit, p.status
            ])
            
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"products_export_{datetime.now().strftime('%Y%m%d')}.csv"
        )
        
    elif format_type == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Products Inventory"
        
        # Headers
        ws.append(['Product Code', 'Product Name', 'Barcode', 'Category', 'Brand', 'Purchase Price', 'Selling Price', 'Current Stock', 'Minimum Stock', 'Maximum Stock', 'Unit', 'Status'])
        
        for p in products:
            ws.append([
                p.product_code, p.product_name, p.barcode, p.category.name, p.brand.brand_name,
                float(p.purchase_price), float(p.selling_price), p.current_stock, p.minimum_stock, p.maximum_stock, p.unit, p.status
            ])
            
        out_stream = io.BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        return send_file(
            out_stream,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"products_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
    flash("Invalid export format requested.", "danger")
    return redirect(url_for('products.index'))

@products_bp.route('/import', methods=['POST'])
@login_required
@role_required(['Super Admin', 'Admin'])
def import_data():
    """
    Imports product list from CSV or Excel sheets templates.
    """
    file = request.files.get('import_file')
    if not file or not file.filename:
        flash("Please upload a file to import.", "danger")
        return redirect(url_for('products.index'))
        
    ext = file.filename.rsplit('.', 1)[-1].lower()
    import_count = 0
    
    # Fetch default category/brand fallback
    def_cat = Category.query.first()
    def_brand = Brand.query.first()
    if not def_cat or not def_brand:
        flash("Please seed or configure at least one Category and one Brand before importing.", "danger")
        return redirect(url_for('products.index'))
        
    try:
        if ext == 'csv':
            stream = io.StringIO(file.stream.read().decode("utf-8"), newline=None)
            reader = csv.reader(stream)
            next(reader, None) # Skip header
            
            for row in reader:
                if len(row) < 5 or not row[1]:
                    continue
                    
                sku = row[2].strip() if row[2] else f"BAR-{uuid.uuid4().hex[:8].upper()}"
                if Product.query.filter_by(barcode=sku, deleted_at=None).first():
                    continue # Skip duplicate barcodes
                    
                code = generate_product_code()
                stock = int(row[7]) if len(row) > 7 and row[7] else 0
                
                prod = Product(
                    product_code=code,
                    barcode=sku,
                    qr_code=code,
                    product_name=row[1].strip(),
                    category_id=def_cat.id,
                    brand_id=def_brand.id,
                    purchase_price=decimal.Decimal(row[5].strip() if len(row) > 5 and row[5] else 1.0),
                    selling_price=decimal.Decimal(row[6].strip() if len(row) > 6 and row[6] else 2.0),
                    current_stock=stock,
                    created_by=current_user.id
                )
                db.session.add(prod)
                db.session.flush()
                
                if stock > 0:
                    mov = InventoryMovement(
                        product_id=prod.id,
                        movement_type='Stock In',
                        quantity=stock,
                        previous_stock=0,
                        new_stock=stock,
                        remarks='Import list opening balance.',
                        created_by=current_user.id
                    )
                    db.session.add(mov)
                    
                import_count += 1
                
        elif ext in ['xlsx', 'xls']:
            wb = load_workbook(file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                for row in rows[1:]:
                    if not row or len(row) < 4 or not row[1]:
                        continue
                        
                    sku = str(row[2]).strip() if row[2] else f"BAR-{uuid.uuid4().hex[:8].upper()}"
                    if Product.query.filter_by(barcode=sku, deleted_at=None).first():
                        continue
                        
                    code = generate_product_code()
                    stock = int(row[7]) if len(row) > 7 and row[7] is not None else 0
                    
                    prod = Product(
                        product_code=code,
                        barcode=sku,
                        qr_code=code,
                        product_name=str(row[1]).strip(),
                        category_id=def_cat.id,
                        brand_id=def_brand.id,
                        purchase_price=decimal.Decimal(row[5] if len(row) > 5 and row[5] is not None else 1.0),
                        selling_price=decimal.Decimal(row[6] if len(row) > 6 and row[6] is not None else 2.0),
                        current_stock=stock,
                        created_by=current_user.id
                    )
                    db.session.add(prod)
                    db.session.flush()
                    
                    if stock > 0:
                        mov = InventoryMovement(
                            product_id=prod.id,
                            movement_type='Stock In',
                            quantity=stock,
                            previous_stock=0,
                            new_stock=stock,
                            remarks='Import list opening balance.',
                            created_by=current_user.id
                        )
                        db.session.add(mov)
                        
                    import_count += 1
                    
        if import_count > 0:
            db.session.commit()
            flash(f"Imported {import_count} product records successfully.", "success")
        else:
            flash("No new products imported (all records matched existing barcodes).", "warning")
            
    except Exception as e:
        db.session.rollback()
        flash(f"Import failed with error: {str(e)}", "danger")
        
    return redirect(url_for('products.index'))
